"""
Реализация генератора результатов.

Конкретная реализация интерфейса IResultGenerator для создания
Excel файлов с результатами расчета стоимости доставки.

Принципы SOLID:
- Single Responsibility Principle (SRP): Отвечает только за генерацию файлов
- Open/Closed Principle (OCP): Легко добавить поддержку новых форматов
- Dependency Inversion Principle (DIP): Реализует абстракцию
"""
import pandas as pd
import logging
import tempfile
from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path

from ..interfaces.i_result_generator import IResultGenerator


logger = logging.getLogger(__name__)


class ExcelResultGenerator(IResultGenerator):
    """
    Конкретная реализация генератора Excel файлов с результатами.
    
    Создает форматированные Excel файлы с результатами расчета
    стоимости доставки, включая детальную информацию и сводки.
    """
    
    def __init__(self):
        """
        Инициализирует генератор результатов.
        """
        # Поддерживаемые форматы файлов
        self._supported_formats = ['xlsx', 'csv', 'json']
        
        # Стили для Excel файла
        self._excel_styles = {
            'header': {
                'font': {'bold': True, 'color': 'FFFFFF'},
                'fill': {'start_color': '366092', 'end_color': '366092', 'fill_type': 'solid'},
                'border': {'style': 'thin'},
                'alignment': {'horizontal': 'center', 'vertical': 'center'}
            },
            'data': {
                'border': {'style': 'thin'},
                'alignment': {'horizontal': 'left', 'vertical': 'center'}
            },
            'price': {
                'number_format': '#,##0.00 ₽',
                'border': {'style': 'thin'},
                'alignment': {'horizontal': 'right', 'vertical': 'center'}
            }
        }
        
        logger.info("ExcelResultGenerator инициализирован")
    
    def generate_result_file(
        self, 
        calculation_results: List[Dict[str, Any]],
        output_format: str = "xlsx"
    ) -> str:
        """
        Создает файл с результатами расчетов стоимости доставки.
        
        Args:
            calculation_results (List[Dict[str, Any]]): Результаты расчетов
            output_format (str): Формат выходного файла
            
        Returns:
            str: Путь к созданному файлу
        """
        if output_format not in self._supported_formats:
            raise ValueError(f"Неподдерживаемый формат: {output_format}")
        
        if not calculation_results:
            raise ValueError("Нет данных для создания файла результатов")
        
        logger.info(f"Генерирую файл результатов в формате {output_format}")
        logger.info(f"Обрабатываю {len(calculation_results)} результатов расчетов")
        
        # Создаем временный файл
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_file = tempfile.NamedTemporaryFile(
            delete=False, 
            suffix=f'.{output_format}',
            prefix=f'shipping_results_{timestamp}_'
        )
        temp_file.close()
        
        try:
            # Основное создание файла
            if output_format == 'xlsx':
                self._generate_excel_file(calculation_results, temp_file.name)
            elif output_format == 'csv':
                self._generate_csv_file(calculation_results, temp_file.name)
            elif output_format == 'json':
                self._generate_json_file(calculation_results, temp_file.name)
            
            # Проверяем, что файл действительно создан
            if not Path(temp_file.name).exists():
                raise ValueError("Файл не был создан")
                
            file_size = Path(temp_file.name).stat().st_size
            if file_size == 0:
                raise ValueError("Создан пустой файл")
            
            logger.info(f"Файл результатов создан: {temp_file.name} (размер: {file_size} байт)")
            return temp_file.name
            
        except Exception as e:
            # Удаляем временный файл при ошибке
            try:
                Path(temp_file.name).unlink(missing_ok=True)
                logger.debug(f"Удален поврежденный временный файл: {temp_file.name}")
            except Exception as cleanup_error:
                logger.warning(f"Не удалось удалить поврежденный файл: {cleanup_error}")
            
            logger.error(f"Ошибка создания файла результатов: {e}")
            
            # Пытаемся создать CSV как fallback
            if output_format == 'xlsx':
                try:
                    logger.info("Пытаюсь создать CSV файл как резервный вариант...")
                    return self.generate_result_file(calculation_results, 'csv')
                except Exception as csv_error:
                    logger.error(f"Не удалось создать резервный CSV файл: {csv_error}")
            
            raise ValueError(f"Не удалось создать файл результатов: {e}")
    
    def get_supported_formats(self) -> List[str]:
        """
        Возвращает список поддерживаемых форматов файлов.
        
        Returns:
            List[str]: Список поддерживаемых форматов
        """
        return self._supported_formats.copy()
    
    def create_summary_sheet(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Создает сводку результатов для всех маршрутов.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            
        Returns:
            Dict[str, Any]: Сводная информация
        """
        if not results:
            return {
                'total_routes': 0,
                'successful_calculations': 0,
                'average_costs': {},
                'popular_companies': {}
            }
        
        total_routes = len(results)
        successful_calculations = sum(1 for r in results if r.get('is_successful', False))
        
        # Собираем все предложения для анализа
        all_offers = []
        for result in results:
            weight_results = result.get('weight_results', {})
            for weight_data in weight_results.values():
                offers = weight_data.get('offers', [])
                all_offers.extend(offers)
        
        # Рассчитываем средние стоимости по весам
        weight_costs = {}
        for offer in all_offers:
            weight = offer.get('weight', 0)
            price = offer.get('price', 0)
            
            if weight not in weight_costs:
                weight_costs[weight] = []
            weight_costs[weight].append(price)
        
        average_costs = {
            weight: sum(prices) / len(prices) 
            for weight, prices in weight_costs.items()
        }
        
        # Находим популярные компании
        company_counts = {}
        for offer in all_offers:
            company = offer.get('company_name', 'Неизвестная')
            company_counts[company] = company_counts.get(company, 0) + 1
        
        # Сортируем компании по популярности
        popular_companies = dict(
            sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        )
        
        summary = {
            'total_routes': total_routes,
            'successful_calculations': successful_calculations,
            'success_rate': (successful_calculations / total_routes * 100) if total_routes > 0 else 0,
            'total_offers': len(all_offers),
            'average_costs': average_costs,
            'popular_companies': popular_companies,
            'unique_companies': len(company_counts),
            'generation_time': datetime.now().isoformat()
        }
        
        logger.info(f"Создана сводка: {successful_calculations}/{total_routes} успешных расчетов")
        return summary
    
    def _generate_excel_file(self, results: List[Dict[str, Any]], file_path: str) -> None:
        """
        Генерирует простой Excel файл с результатами используя прямое openpyxl.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            file_path (str): Путь к файлу для создания
        """
        try:
            from openpyxl import Workbook
            
            # Подготавливаем простые данные
            simple_data = self._prepare_simple_data(results)
            
            if not simple_data:
                raise ValueError("Нет данных для создания файла")
            
            logger.info(f"Создаю Excel файл с {len(simple_data)} строками данных")
            
            # Создаем новый Excel файл напрямую через openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты"
            
            # Добавляем заголовки
            headers = ['Маршрут', 'Вес_кг', 'Компания', 'Цена_руб', 'Срок_дней']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Добавляем данные
            for row_num, row_data in enumerate(simple_data, 2):
                ws.cell(row=row_num, column=1, value=row_data.get('Маршрут', ''))
                ws.cell(row=row_num, column=2, value=row_data.get('Вес_кг', 0.0))
                ws.cell(row=row_num, column=3, value=row_data.get('Компания', ''))
                ws.cell(row=row_num, column=4, value=row_data.get('Цена_руб', 0.0))
                ws.cell(row=row_num, column=5, value=row_data.get('Срок_дней', 0))
            
            # Сохраняем файл
            wb.save(file_path)
            wb.close()
            
            # Проверяем размер файла
            file_size = Path(file_path).stat().st_size
            if file_size < 1000:  # Слишком маленький файл
                raise ValueError(f"Созданный файл слишком мал: {file_size} байт")
            
            logger.info(f"Excel файл успешно создан с {len(simple_data)} строками (размер: {file_size} байт)")
            
        except Exception as e:
            logger.error(f"Ошибка создания Excel файла через openpyxl: {e}")
            raise ValueError(f"Не удалось создать Excel файл: {e}")
    
    def _generate_csv_file(self, results: List[Dict[str, Any]], file_path: str) -> None:
        """
        Генерирует CSV файл с результатами.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов  
            file_path (str): Путь к файлу для создания
        """
        main_data = self._prepare_main_data(results)
        df = pd.DataFrame(main_data)
        df.to_csv(file_path, index=False, encoding='utf-8-sig')
        logger.info("CSV файл создан")
    
    def _generate_json_file(self, results: List[Dict[str, Any]], file_path: str) -> None:
        """
        Генерирует JSON файл с результатами.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            file_path (str): Путь к файлу для создания
        """
        import json
        
        output_data = {
            'summary': self.create_summary_sheet(results),
            'results': results,
            'generation_info': {
                'timestamp': datetime.now().isoformat(),
                'total_routes': len(results),
                'generator': 'ExcelResultGenerator'
            }
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        
        logger.info("JSON файл создан")
    
    def _prepare_simple_data(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Подготавливает простые данные для Excel файла.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            
        Returns:
            List[Dict[str, Any]]: Простые данные для Excel
        """
        simple_data = []
        
        if not results:
            logger.warning("Нет результатов для подготовки простых данных")
            return simple_data
        
        for result in results:
            try:
                route_info = result.get('route', {})
                weight_results = result.get('weight_results', {})
                
                origin = self._safe_convert_str(route_info.get('origin', 'Не указано'))
                destination = self._safe_convert_str(route_info.get('destination', 'Не указано'))
                route_name = f"{origin} → {destination}"
                
                if not weight_results:
                    # Если нет результатов, добавляем строку с ошибкой
                    simple_data.append({
                        'Маршрут': route_name,
                        'Вес_кг': 0.0,
                        'Компания': 'Ошибка расчета',
                        'Цена_руб': 0.0,
                        'Срок_дней': 0
                    })
                    continue
                
                # Создаем строку для каждого веса
                for weight, weight_data in weight_results.items():
                    try:
                        weight_kg = self._safe_convert_float(weight) / 1000 if isinstance(weight, (int, float)) else 0.0
                        cheapest_offer = weight_data.get('cheapest_offer')
                        
                        if cheapest_offer:
                            company_name = self._safe_convert_str(cheapest_offer.get('company_name', 'Неизвестно'))
                            price = self._safe_convert_float(cheapest_offer.get('price', 0))
                            delivery_days = self._safe_convert_int(cheapest_offer.get('delivery_days', 0))
                        else:
                            company_name = 'Нет предложений'
                            price = 0.0
                            delivery_days = 0
                        
                        simple_data.append({
                            'Маршрут': route_name,
                            'Вес_кг': weight_kg,
                            'Компания': company_name,
                            'Цена_руб': price,
                            'Срок_дней': delivery_days
                        })
                        
                    except Exception as e:
                        logger.error(f"Ошибка обработки веса {weight}: {e}")
                        continue
                        
            except Exception as e:
                logger.error(f"Ошибка обработки результата: {e}")
                continue
        
        logger.info(f"Подготовлено {len(simple_data)} простых строк данных")
        return simple_data
    
    def _prepare_main_data(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Подготавливает основные данные для файла результатов с валидацией.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            
        Returns:
            List[Dict[str, Any]]: Подготовленные и валидированные данные
        """
        main_data = []
        
        if not results:
            logger.warning("Нет результатов для подготовки данных")
            return main_data
        
        for result in results:
            try:
                route_info = result.get('route', {})
                weight_results = result.get('weight_results', {})
                
                if not weight_results:
                    logger.debug(f"Нет весовых результатов для маршрута: {route_info}")
                    continue
                
                # Создаем строку для каждого веса
                for weight, weight_data in weight_results.items():
                    try:
                        cheapest_offer = weight_data.get('cheapest_offer')
                        
                        # Валидация и очистка данных
                        row_index = self._safe_convert_int(route_info.get('row_index', 0))
                        origin = self._safe_convert_str(route_info.get('origin', 'Не указано'))
                        destination = self._safe_convert_str(route_info.get('destination', 'Не указано'))
                        weight_kg = self._safe_convert_float(weight) / 1000 if isinstance(weight, (int, float)) else 0.0
                        
                        # Данные предложения с валидацией
                        if cheapest_offer:
                            company_name = self._safe_convert_str(cheapest_offer.get('company_name', 'Неизвестно'))
                            price = self._safe_convert_float(cheapest_offer.get('price', 0))
                            delivery_days = self._safe_convert_int(cheapest_offer.get('delivery_days', 0))
                            tariff_name = self._safe_convert_str(cheapest_offer.get('tariff_name', ''))
                            price_per_kg = self._safe_convert_float(cheapest_offer.get('price_per_kg', 0))
                            status = 'Успешно'
                        else:
                            company_name = ''
                            price = 0.0
                            delivery_days = 0
                            tariff_name = ''
                            price_per_kg = 0.0
                            status = 'Ошибка'
                        
                        offers_count = self._safe_convert_int(weight_data.get('offers_count', 0))
                        
                        row = {
                            'Номер_строки': row_index,
                            'Откуда': origin,
                            'Куда': destination,
                            'Вес_кг': weight_kg,
                            'Статус': status,
                            'Компания': company_name,
                            'Цена_руб': price,
                            'Срок_дней': delivery_days,
                            'Тариф': tariff_name,
                            'Цена_за_кг': price_per_kg,
                            'Количество_предложений': offers_count
                        }
                        
                        main_data.append(row)
                        
                    except Exception as e:
                        logger.error(f"Ошибка обработки веса {weight} для маршрута {route_info}: {e}")
                        continue
                        
            except Exception as e:
                logger.error(f"Ошибка обработки результата: {e}")
                continue
        
        # Сортируем по номеру строки и весу
        try:
            main_data.sort(key=lambda x: (x.get('Номер_строки', 0), x.get('Вес_кг', 0)))
        except Exception as e:
            logger.error(f"Ошибка сортировки данных: {e}")
        
        logger.info(f"Подготовлено {len(main_data)} строк данных")
        return main_data
    
    def _prepare_weight_data(
        self, 
        results: List[Dict[str, Any]], 
        target_weight: int
    ) -> List[Dict[str, Any]]:
        """
        Подготавливает данные для конкретной весовой категории.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            target_weight (int): Целевой вес в граммах
            
        Returns:
            List[Dict[str, Any]]: Данные для веса
        """
        weight_data = []
        
        for result in results:
            route_info = result.get('route', {})
            weight_results = result.get('weight_results', {})
            
            if target_weight in weight_results:
                weight_result = weight_results[target_weight]
                offers = weight_result.get('offers', [])
                
                for offer in offers:
                    row = {
                        'Откуда': route_info.get('origin', ''),
                        'Куда': route_info.get('destination', ''),
                        'Компания': offer.get('company_name', ''),
                        'Цена_руб': offer.get('price', 0),
                        'Срок_дней': offer.get('delivery_days', 0),
                        'Тариф': offer.get('tariff_name', ''),
                        'Цена_за_кг': offer.get('price_per_kg', 0)
                    }
                    weight_data.append(row)
        
        # Сортируем по цене
        weight_data.sort(key=lambda x: x.get('Цена_руб', 0))
        
        return weight_data
    
    def _get_weight_categories(self, results: List[Dict[str, Any]]) -> List[int]:
        """
        Извлекает все использованные весовые категории.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            
        Returns:
            List[int]: Список весов в граммах
        """
        weights = set()
        
        for result in results:
            weight_results = result.get('weight_results', {})
            weights.update(weight_results.keys())
        
        return sorted(list(weights))
    
    def _safe_convert_str(self, value) -> str:
        """
        Безопасно конвертирует значение в строку.
        
        Args:
            value: Значение для конвертации
            
        Returns:
            str: Строковое представление значения
        """
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        try:
            return str(value).strip()
        except Exception:
            return ''
    
    def _safe_convert_float(self, value) -> float:
        """
        Безопасно конвертирует значение в float.
        
        Args:
            value: Значение для конвертации
            
        Returns:
            float: Числовое значение или 0.0
        """
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(',', '.').replace(' ', ''))
            except (ValueError, AttributeError):
                return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _safe_convert_int(self, value) -> int:
        """
        Безопасно конвертирует значение в int.
        
        Args:
            value: Значение для конвертации
            
        Returns:
            int: Целое число или 0
        """
        if value is None:
            return 0
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            try:
                return int(float(value.replace(',', '.').replace(' ', '')))
            except (ValueError, AttributeError):
                return 0
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0