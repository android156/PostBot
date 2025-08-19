"""
Сервис Telegram бота.

Основная бизнес-логика для работы с Telegram ботом,
обработки команд пользователей и координации работы всех компонентов.

Принципы SOLID:
- Single Responsibility Principle (SRP): Отвечает только за логику бота
- Open/Closed Principle (OCP): Легко расширяется новыми командами
- Dependency Inversion Principle (DIP): Зависит от абстракций, а не от конкретных классов
"""
import logging
import tempfile
import asyncio
from typing import Dict, Any, List
from datetime import datetime

from telegram import Update, InputFile
from telegram.ext import ContextTypes
from telegram.constants import ParseMode

from ..interfaces.i_bot_service import IBotService
from ..interfaces.i_excel_processor import IExcelProcessor
from ..interfaces.i_api_client import IApiClient
from ..interfaces.i_result_generator import IResultGenerator
from ..interfaces.i_config import IConfig
from ..models.route import Route
from ..models.shipping_calculation import (
    ShippingOffer, WeightCategoryResult, RouteCalculationResult
)


logger = logging.getLogger(__name__)


class BotService(IBotService):
    """
    Конкретная реализация сервиса Telegram бота.
    
    Координирует работу всех компонентов системы:
    - Обрабатывает команды и сообщения пользователей
    - Управляет обработкой Excel файлов
    - Выполняет расчеты стоимости доставки
    - Генерирует файлы с результатами
    
    Использует внедрение зависимостей для всех компонентов.
    """
    
    def __init__(
        self,
        config: IConfig,
        excel_processor: IExcelProcessor,
        api_client: IApiClient,
        result_generator: IResultGenerator
    ):
        """
        Инициализирует сервис бота с внедрением зависимостей.
        
        Args:
            config (IConfig): Интерфейс конфигурации
            excel_processor (IExcelProcessor): Процессор Excel файлов
            api_client (IApiClient): Клиент API доставки
            result_generator (IResultGenerator): Генератор результатов
        """
        self._config = config
        self._excel_processor = excel_processor
        self._api_client = api_client
        self._result_generator = result_generator
        
        # Получаем настройки из конфигурации
        self._file_settings = self._config.get_file_processing_settings()
        self._weight_categories = self._config.get_weight_categories()
        
        # Статистика для мониторинга
        self._stats = {
            'total_files_processed': 0,
            'total_routes_calculated': 0,
            'total_api_calls': 0,
            'start_time': datetime.now()
        }
        
        logger.info("BotService инициализирован с внедренными зависимостями")
    
    async def handle_start_command(self, update, context) -> None:
        """
        Обрабатывает команду /start.
        
        Отправляет приветственное сообщение с инструкциями по использованию.
        
        Args:
            update: Объект обновления от Telegram
            context: Контекст бота
        """
        try:
            user_id = update.effective_user.id if update.effective_user else "unknown"
            logger.info(f"Пользователь {user_id} выполнил команду /start")
            
            welcome_message = self._create_welcome_message()
            
            await update.message.reply_text(welcome_message, parse_mode=ParseMode.MARKDOWN)
            
            logger.info(f"Отправлено приветственное сообщение пользователю {user_id}")
            
        except Exception as e:
            logger.error(f"Ошибка обработки команды /start: {e}")
            # await self._send_error_message(update, "Ошибка при запуске бота")
    
    async def handle_help_command(self, update, context) -> None:
        """
        Обрабатывает команду /help.
        
        Отправляет подробную справку по использованию бота.
        
        Args:
            update: Объект обновления от Telegram
            context: Контекст бота
        """
        try:
            user_id = update.effective_user.id if update.effective_user else "unknown"
            logger.info(f"Пользователь {user_id} запросил справку")
            
            help_message = self._create_help_message()
            
            await update.message.reply_text(help_message, parse_mode=ParseMode.MARKDOWN)
            
        except Exception as e:
            logger.error(f"Ошибка обработки команды /help: {e}")
    
    async def handle_document(self, update, context) -> None:
        """
        Обрабатывает загруженные документы (Excel файлы).
        
        Основной рабочий процесс:
        1. Проверяет файл
        2. Скачивает и обрабатывает
        3. Рассчитывает стоимость доставки
        4. Создает и отправляет файл с результатами
        
        Args:
            update: Объект обновления от Telegram с документом
            context: Контекст бота
        """
        document = update.message.document if update.message else None
        user_id = update.effective_user.id if update.effective_user else "unknown"
        
        if not document:
            logger.error("Нет документа в сообщении")
            return
            
        logger.info(f"Пользователь {user_id} загрузил файл: {document.file_name if document.file_name else 'без имени'}")
        
        try:
            # Шаг 1: Валидация файла
            validation_result = await self._validate_uploaded_file(document)
            if not validation_result['valid']:
                # await update.message.reply_text(validation_result['error_message'])
                return
            
            # Шаг 2: Уведомление о начале обработки
            processing_msg = await update.message.reply_text("📊 Обрабатываю файл...")
            
            # Шаг 3: Скачивание и обработка файла
            routes_data = await self._download_and_process_file(document, context)
            if not routes_data:
                await processing_msg.edit_text("❌ Не удалось извлечь данные из файла")
                return
                
            await processing_msg.edit_text(
                f"📋 Найдено {len(routes_data)} маршрутов. Рассчитываю стоимость доставки..."
            )
            
            # Шаг 4: Расчет стоимости доставки
            calculation_results = await self.process_shipping_calculation(routes_data)
            
            # Шаг 5: Генерация файла с результатами
            result_file_path = await self._generate_result_file(calculation_results)
            
            # Шаг 6: Отправка результатов пользователю
            await self._send_results_to_user(update, result_file_path, calculation_results)
            
            # Обновляем статистику
            self._stats['total_files_processed'] += 1
            self._stats['total_routes_calculated'] += len(routes_data)
            
            logger.info(f"Успешно обработан файл от пользователя {user_id}")
            
        except Exception as e:
            logger.error(f"Ошибка обработки файла от пользователя {user_id}: {e}")
            # await self._send_error_message(update, f"Ошибка обработки файла: {str(e)}")
    
    async def handle_text_message(self, update, context) -> None:
        """
        Обрабатывает текстовые сообщения от пользователей.
        
        Args:
            update: Объект обновления от Telegram с текстом
            context: Контекст бота
        """
        try:
            user_id = update.effective_user.id if update.effective_user else "unknown"
            message_text = update.message.text if update.message else None
            
            if not message_text:
                logger.error("Нет текста в сообщении")
                return
                
            logger.info(f"Получено текстовое сообщение от пользователя {user_id}: {message_text}")
            
            # Анализируем сообщение и отвечаем соответствующе
            response = self._analyze_text_message(message_text)
            
            await update.message.reply_text(response)
            
        except Exception as e:
            logger.error(f"Ошибка обработки текстового сообщения: {e}")
    
    async def process_shipping_calculation(
        self, 
        routes_data: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """
        Выполняет расчет стоимости доставки для списка маршрутов.
        
        Основной алгоритм расчета:
        1. Предварительно получает коды городов для всех маршрутов
        2. Для каждого маршрута и весовой категории выполняет расчет
        3. Находит лучшие предложения
        4. Формирует результат
        
        Args:
            routes_data (List[Dict[str, Any]]): Список маршрутов для расчета
            
        Returns:
            Dict[str, Any]: Результаты расчета для всех маршрутов
        """
        logger.info(f"Начинаю расчет стоимости для {len(routes_data)} маршрутов")
        logger.info(f"Весовые категории: {self._weight_categories}")
        
        # Шаг 1: Предварительно получаем коды городов для всех маршрутов
        routes_with_codes = await self._resolve_all_city_codes(routes_data)
        
        calculation_results = []
        total_calculations = len(routes_with_codes) * len(self._weight_categories)
        completed_calculations = 0
        
        for route_data in routes_with_codes:
            try:
                # Создаем объект маршрута
                route = Route(
                    origin=route_data['origin'],
                    destination=route_data['destination'],
                    row_index=route_data['row_index']
                )
                
                # Проверяем, что коды городов найдены
                origin_code = route_data.get('origin_code')
                destination_code = route_data.get('destination_code')
                
                if not origin_code or not destination_code:
                    logger.warning(f"Пропускаю маршрут {route.get_display_name()} - не найдены коды городов")
                    continue
                
                # Создаем результат для маршрута
                route_result = RouteCalculationResult(route=route)
                
                # Рассчитываем для каждой весовой категории
                for weight in self._weight_categories:
                    try:
                        logger.info(f"Расчет для {route.get_display_name()}, вес {weight}кг")
                        
                        # Используем оптимизированный метод прямого расчета с готовыми кодами
                        api_result = await self._api_client.calculate_shipping_cost_with_codes(
                            origin_code, destination_code, weight
                        )
                        
                        logger.info(f"Результат API для {route.get_display_name()}: success={api_result.get('success', False)}")
                        
                        # Создаем результат для весовой категории
                        weight_result = self._process_api_result(api_result, weight)
                        route_result.add_weight_result(weight_result)
                        
                        completed_calculations += 1
                        self._stats['total_api_calls'] += 1
                        
                        # Логируем прогресс каждые 10%
                        progress = (completed_calculations / total_calculations) * 100
                        if completed_calculations % max(1, total_calculations // 10) == 0:
                            logger.info(f"Прогресс расчета: {progress:.1f}% ({completed_calculations}/{total_calculations})")
                        
                    except Exception as e:
                        logger.error(f"Ошибка расчета для {route.get_display_name()}, вес {weight}кг: {e}")
                        # Создаем результат с ошибкой
                        error_result = WeightCategoryResult(weight=weight, calculation_error=str(e))
                        route_result.add_weight_result(error_result)
                
                calculation_results.append(route_result.to_dict())
                
            except Exception as e:
                logger.error(f"Ошибка обработки маршрута {route_data}: {e}")
                continue
        
        # Создаем общую сводку результатов
        summary = self._create_calculation_summary(calculation_results)
        
        result = {
            'success': True,
            'results': calculation_results,
            'summary': summary,
            'total_routes': len(routes_data),
            'processed_routes': len(calculation_results),
            'total_api_calls': self._stats['total_api_calls'],
            'calculation_time': datetime.now().isoformat()
        }
        
        logger.info(f"Расчет завершен: {len(calculation_results)} маршрутов обработано")
        return result
    
    async def start_bot(self) -> None:
        """
        Запускает бота и начинает опрос обновлений.
        
        Инициализирует все компоненты и начинает основной цикл работы.
        """
        logger.info("Запуск сервиса Telegram бота...")
        
        try:
            # Проверяем конфигурацию
            if not self._config.validate_configuration():
                raise ValueError("Неверная конфигурация приложения")
            
            # Инициализируем API клиент
            if not await self._api_client.authenticate():
                logger.error("Не удалось аутентифицироваться в API доставки")
                raise ValueError("Ошибка аутентификации API")
            
            logger.info("Сервис бота успешно запущен и готов к работе")
            
        except Exception as e:
            logger.error(f"Ошибка запуска сервиса бота: {e}")
            raise
    
    async def stop_bot(self) -> None:
        """
        Останавливает бота и освобождает ресурсы.
        """
        logger.info("Остановка сервиса Telegram бота...")
        
        try:
            # Закрываем API клиент
            await self._api_client.close()
            
            # Логируем статистику
            runtime = datetime.now() - self._stats['start_time']
            logger.info(f"Статистика работы:")
            logger.info(f"- Время работы: {runtime}")
            logger.info(f"- Обработано файлов: {self._stats['total_files_processed']}")
            logger.info(f"- Рассчитано маршрутов: {self._stats['total_routes_calculated']}")
            logger.info(f"- Вызовов API: {self._stats['total_api_calls']}")
            
            logger.info("Сервис бота успешно остановлен")
            
        except Exception as e:
            logger.error(f"Ошибка остановки сервиса бота: {e}")
    
    # Приватные методы для внутренней логики
    
    def _create_welcome_message(self) -> str:
        """
        Создает приветственное сообщение для команды /start.
        
        Returns:
            str: Отформатированное приветственное сообщение
        """
        weight_categories = [f"{w:.1f}кг" for w in self._weight_categories]
        weights_text = ", ".join(weight_categories)
        
        return f"""
🚚 **Добро пожаловать в бот расчета стоимости доставки!**

Этот бот поможет вам найти самые выгодные варианты доставки для разных весовых категорий.

**Как использовать:**
1. Отправьте Excel файл с маршрутами доставки
2. Файл должен содержать колонки: "Откуда", "Куда" (без указания веса)
3. Бот автоматически протестирует {len(self._weight_categories)} весовых категорий
4. Получите Excel файл с лучшими предложениями для каждого веса

**Тестируемые весовые категории:** {weights_text}

**Максимальный размер файла:** {self._file_settings['max_file_size'] // (1024*1024)}MB

Для получения подробной справки используйте команду /help
        """.strip()
    
    def _create_help_message(self) -> str:
        """
        Создает справочное сообщение для команды /help.
        
        Returns:
            str: Отформатированное справочное сообщение
        """
        supported_columns = self._excel_processor.get_supported_column_names()
        origin_examples = ", ".join(f'"{col}"' for col in supported_columns['origin'][:5])
        destination_examples = ", ".join(f'"{col}"' for col in supported_columns['destination'][:5])
        
        return f"""
📋 **Подробная справка по использованию бота**

**Требования к Excel файлу:**
• Форматы: .xlsx, .xls
• Максимальный размер: {self._file_settings['max_file_size'] // (1024*1024)}MB
• Первая строка должна содержать заголовки колонок
• Обязательно наличие колонок с городами отправления и назначения

**Поддерживаемые названия колонок:**
• Отправление: {origin_examples} и др.
• Назначение: {destination_examples} и др.

**Принцип работы:**
1. Бот читает маршруты из вашего файла
2. Для каждого маршрута выполняет расчет по {len(self._weight_categories)} весовым категориям
3. Находит самые выгодные предложения от транспортных компаний
4. Создает детальный Excel отчет с результатами

**Структура результатов:**
• Основной лист с лучшими предложениями для каждого веса
• Отдельные листы для каждой весовой категории
• Сводный лист с общей статистикой

**Важные особенности:**
• Не указывайте вес в исходном файле - он тестируется автоматически
• Бот поддерживает как названия городов, так и коды КЛАДР
• Результаты сортируются по выгодности предложений

При возникновении проблем проверьте формат файла и названия колонок.
        """.strip()
    
    def _analyze_text_message(self, message_text: str) -> str:
        """
        Анализирует текстовое сообщение и формирует ответ.
        
        Args:
            message_text (str): Текст сообщения от пользователя
            
        Returns:
            str: Ответ пользователю
        """
        message_lower = message_text.lower().strip()
        
        # Определяем тип запроса и отвечаем соответствующе
        if any(word in message_lower for word in ['привет', 'здравствуй', 'добрый день']):
            return "Привет! Отправьте Excel файл с маршрутами доставки для расчета стоимости. Используйте /help для получения справки."
        
        elif any(word in message_lower for word in ['помощь', 'справка', 'как']):
            return "Используйте команду /help для получения подробной справки по работе с ботом."
        
        elif any(word in message_lower for word in ['статус', 'работа', 'доступен']):
            uptime = datetime.now() - self._stats['start_time']
            return f"Бот работает нормально. Время работы: {uptime}. Обработано файлов: {self._stats['total_files_processed']}."
        
        else:
            return ("Отправьте Excel файл с маршрутами доставки для расчета стоимости. "
                   "Для получения справки используйте команду /help.")
    
    async def _validate_uploaded_file(self, document) -> Dict[str, Any]:
        """
        Валидирует загруженный пользователем файл.
        
        Args:
            document: Объект документа от Telegram
            
        Returns:
            Dict[str, Any]: Результат валидации
        """
        # Проверяем размер файла
        if document.file_size > self._file_settings['max_file_size']:
            max_size_mb = self._file_settings['max_file_size'] // (1024*1024)
            return {
                'valid': False,
                'error_message': f"❌ Файл слишком большой. Максимальный размер: {max_size_mb}MB"
            }
        
        # Проверяем расширение файла
        if not document.file_name.lower().endswith(tuple(self._file_settings['allowed_extensions'])):
            extensions = ', '.join(self._file_settings['allowed_extensions'])
            return {
                'valid': False,
                'error_message': f"❌ Поддерживаются только файлы: {extensions}"
            }
        
        return {'valid': True}
    
    async def _download_and_process_file(self, document, context) -> List[Dict[str, Any]]:
        """
        Скачивает и обрабатывает Excel файл.
        
        Args:
            document: Объект документа от Telegram
            context: Контекст бота
            
        Returns:
            List[Dict[str, Any]]: Данные маршрутов из файла
        """
        try:
            # Скачиваем файл
            file = await context.bot.get_file(document.file_id)
            
            # Создаем временный файл
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                await file.download_to_drive(tmp_file.name)
                temp_file_path = tmp_file.name
            
            try:
                # Обрабатываем Excel файл
                routes_data = self._excel_processor.process_file(temp_file_path)
                logger.info(f"Извлечено {len(routes_data)} маршрутов из файла")
                return routes_data
                
            finally:
                # Удаляем временный файл
                import os
                os.unlink(temp_file_path)
                
        except Exception as e:
            logger.error(f"Ошибка скачивания и обработки файла: {e}")
            return []
    
    def _process_api_result(self, api_result: Dict[str, Any], weight: float) -> WeightCategoryResult:
        """
        Обрабатывает результат API в объект WeightCategoryResult.
        
        Args:
            api_result (Dict[str, Any]): Результат от API
            weight (float): Вес категории в килограммах
            
        Returns:
            WeightCategoryResult: Обработанный результат
        """
        # Конвертируем вес в граммы для WeightCategoryResult
        weight_in_grams = int(weight * 1000)
        weight_result = WeightCategoryResult(weight=weight_in_grams)
        
        if api_result.get('success'):
            # Обрабатываем успешный результат
            offers_data = api_result.get('offers', [])
            
            for offer_data in offers_data:
                try:
                    offer = ShippingOffer(
                        company_name=offer_data.get('company_name', ''),
                        price=float(offer_data.get('price', 0)),
                        delivery_days=int(offer_data.get('delivery_days', 0)),
                        tariff_name=offer_data.get('tariff_name', ''),
                        weight=weight_in_grams,  # Используем вес в граммах
                        additional_info=offer_data.get('additional_info', {})
                    )
                    weight_result.add_offer(offer)
                    
                except (ValueError, TypeError) as e:
                    logger.warning(f"Ошибка создания предложения: {e}")
                    continue
        else:
            # Устанавливаем ошибку
            weight_result.calculation_error = api_result.get('error', 'Неизвестная ошибка API')
        
        return weight_result
    
    def _create_calculation_summary(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Создает сводку результатов расчета.
        
        Args:
            results (List[Dict[str, Any]]): Результаты расчетов
            
        Returns:
            Dict[str, Any]: Сводная информация
        """
        successful_routes = sum(1 for r in results if r.get('is_successful', False))
        
        # Собираем статистику по весам
        weight_stats = {}
        for result in results:
            weight_results = result.get('weight_results', {})
            for weight, weight_data in weight_results.items():
                if weight not in weight_stats:
                    weight_stats[weight] = {'successful': 0, 'total': 0}
                
                weight_stats[weight]['total'] += 1
                if weight_data.get('has_offers'):
                    weight_stats[weight]['successful'] += 1
        
        return {
            'total_routes': len(results),
            'successful_routes': successful_routes,
            'success_rate': (successful_routes / len(results) * 100) if results else 0,
            'weight_statistics': weight_stats,
            'processing_time': datetime.now().isoformat()
        }
    
    async def _generate_result_file(self, calculation_results: Dict[str, Any]) -> str:
        """
        Генерирует файл с результатами расчета.
        
        Args:
            calculation_results (Dict[str, Any]): Результаты расчета
            
        Returns:
            str: Путь к сгенерированному файлу
        """
        try:
            results_data = calculation_results.get('results', [])
            result_file_path = self._result_generator.generate_result_file(
                results_data, 
                output_format='xlsx'
            )
            
            logger.info(f"Создан файл результатов: {result_file_path}")
            return result_file_path
            
        except Exception as e:
            logger.error(f"Ошибка генерации файла результатов: {e}")
            raise
    
    async def _send_results_to_user(self, update, result_file_path: str, results: Dict[str, Any]) -> None:
        """
        Отправляет файл с результатами пользователю.
        
        Args:
            update: Объект обновления от Telegram
            result_file_path (str): Путь к файлу результатов
            results (Dict[str, Any]): Данные результатов
        """
        try:
            # Проверяем файл перед отправкой
            if not self._validate_result_file(result_file_path):
                logger.error(f"Файл {result_file_path} поврежден, создаю CSV")
                # Создаем CSV как fallback
                csv_path = await self._create_csv_fallback(results)
                result_file_path = csv_path
            
            summary = results.get('summary', {})
            
            # Формируем текст с кратким отчетом
            report_text = f"""
✅ **Расчет завершен!**

📊 **Статистика:**
• Всего маршрутов: {summary.get('total_routes', 0)}
• Успешных расчетов: {summary.get('successful_routes', 0)}
• Процент успеха: {summary.get('success_rate', 0):.1f}%
• Всего вызовов API: {results.get('total_api_calls', 0)}

📋 Подробные результаты во вложенном Excel файле.
            """.strip()
            
            # Определяем имя файла и MIME-тип
            if result_file_path.endswith('.csv'):
                filename = "shipping_results.csv"
                report_text += "\n\n📝 *Файл отправлен в формате CSV для лучшей совместимости*"
            else:
                filename = "shipping_results.xlsx"
            
            # Отправляем файл с явным указанием параметров
            with open(result_file_path, 'rb') as file:
                await update.message.reply_document(
                    document=file,
                    filename=filename,
                    caption=report_text,
                    parse_mode=ParseMode.MARKDOWN
                )
            
            logger.info("Результаты успешно отправлены пользователю")
            
            # Удаляем временный файл
            try:
                import os
                os.unlink(result_file_path)
                logger.debug(f"Удален временный файл: {result_file_path}")
            except Exception as cleanup_error:
                logger.warning(f"Не удалось удалить временный файл {result_file_path}: {cleanup_error}")
            
        except Exception as e:
            logger.error(f"Ошибка отправки результатов: {e}")
            try:
                await update.message.reply_text("❌ Произошла ошибка при отправке результатов. Попробуйте еще раз.")
            except:
                pass
    
    def _validate_result_file(self, file_path: str) -> bool:
        """
        Проверяет целостность созданного файла результатов.
        
        Args:
            file_path (str): Путь к файлу для проверки
            
        Returns:
            bool: True если файл корректен
        """
        try:
            from pathlib import Path
            
            file_obj = Path(file_path)
            if not file_obj.exists():
                logger.error(f"Файл не существует: {file_path}")
                return False
            
            file_size = file_obj.stat().st_size
            if file_size < 1000:  # Excel файл должен быть больше 1KB
                logger.error(f"Файл слишком мал: {file_size} байт")
                return False
            
            # Проверяем, что это действительно Excel файл
            if file_path.endswith('.xlsx'):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active
                    if ws.max_row < 2:  # Должно быть заголовки + данные
                        logger.error("Excel файл пуст или содержит только заголовки")
                        return False
                    wb.close()
                    logger.info(f"Excel файл прошел валидацию: {file_size} байт, {ws.max_row} строк")
                    return True
                except Exception as e:
                    logger.error(f"Ошибка чтения Excel файла: {e}")
                    return False
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка валидации файла: {e}")
            return False
    
    async def _create_csv_fallback(self, results: Dict[str, Any]) -> str:
        """
        Создает CSV файл как резервный вариант если Excel не работает.
        
        Args:
            results (Dict[str, Any]): Данные результатов
            
        Returns:
            str: Путь к созданному CSV файлу
        """
        try:
            results_data = results.get('results', [])
            csv_path = self._result_generator.generate_result_file(
                results_data, 
                output_format='csv'
            )
            
            logger.info(f"Создан CSV файл как fallback: {csv_path}")
            return csv_path
            
        except Exception as e:
            logger.error(f"Ошибка создания CSV fallback: {e}")
            raise
    
    async def _resolve_all_city_codes(self, routes_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Предварительно получает коды городов для всех маршрутов.
        
        Оптимизирует производительность, получая коды городов один раз
        для каждого маршрута вместо повторных запросов на каждый вес.
        
        Args:
            routes_data (List[Dict[str, Any]]): Исходные данные маршрутов
            
        Returns:
            List[Dict[str, Any]]: Маршруты с добавленными кодами городов
        """
        logger.info("Предварительно получаю коды городов для всех маршрутов...")
        
        # Собираем уникальные города для батчевого запроса
        unique_cities = set()
        for route_data in routes_data:
            unique_cities.add(route_data['origin'])
            unique_cities.add(route_data['destination'])
        
        logger.info(f"Найдено {len(unique_cities)} уникальных городов для резолва")
        
        # Кеш для кодов городов
        city_codes_cache = {}
        
        # Получаем коды для всех уникальных городов
        for city in unique_cities:
            try:
                # Используем существующий метод API клиента
                city_code = await self._api_client._resolve_city_code(city)
                city_codes_cache[city] = city_code
                
                if city_code:
                    logger.debug(f"Получен код для города '{city}': {city_code}")
                else:
                    logger.warning(f"Не найден код для города '{city}'")
                    
            except Exception as e:
                logger.error(f"Ошибка получения кода для города '{city}': {e}")
                city_codes_cache[city] = None
        
        # Обогащаем маршруты кодами городов
        routes_with_codes = []
        for route_data in routes_data:
            enhanced_route = route_data.copy()
            enhanced_route['origin_code'] = city_codes_cache.get(route_data['origin'])
            enhanced_route['destination_code'] = city_codes_cache.get(route_data['destination'])
            routes_with_codes.append(enhanced_route)
        
        successful_routes = sum(1 for r in routes_with_codes 
                              if r.get('origin_code') and r.get('destination_code'))
        
        logger.info(f"Коды городов получены: {successful_routes}/{len(routes_data)} маршрутов готовы к расчету")
        
        return routes_with_codes
    
    # async def _send_error_message(self, update, error_message: str) -> None:
    #     """
    #     Отправляет сообщение об ошибке пользователю.
        
    #     Args:
    #         update: Объект обновления от Telegram
    #         error_message (str): Текст ошибки
    #     """
    #     try:
    #         await update.message.reply_text(f"❌ {error_message}")
    #     except Exception as e:
    #         logger.error(f"Не удалось отправить сообщение об ошибке: {e}")